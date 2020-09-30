import openpyxl as xl
import json
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
import time 
import sys

json_file = 'predicted.json'
template = 'template.xlsx'
save_file = 'output.xlsx'

f = open(json_file)
data = json.load(f)

wb = xl.load_workbook(template)
ws = wb.active
sheet = wb[wb.sheetnames[0]]

border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

seller_state = sheet['e3']
seller_id = sheet['e4']
seller_name = sheet['e5']
seller_address = sheet['e6']
seller_gstin = sheet['e11']
seller_country = sheet['e12']
currency = sheet['e13']
description = sheet['e14']
invoice_number = sheet['m3']
invoice_date = sheet['m4']
due_date = sheet['m5']
total_invoice_amount = sheet['m6']
total_invoice_quantity = sheet['m7']
total_tcs = sheet['m8']
round_off = sheet['m9']
po_number =  sheet['m10']
items_total = sheet['m11']
items_quantity = sheet['m12']
buyer_gstin = sheet['m13']
shipping_address = sheet['m14']

seller_name.value = data["seller_name"]
seller_address.value = data["seller_address"]
seller_gstin.value = data["seller_gstin"]
invoice_number.value = data["invoice_number"]
invoice_date.value = data["invoice_date"]
total_invoice_amount.value = data["net_total"]

if "due_date" in data:
    due_date.value = data["due_date"]
else:
    due_date.value = data["invoice_date"]

if "seller_country" in data:
    seller_country.value = data["seller_country"]
else:
    seller_country.value = "India"

if "currency" in data:
    currency.value = data["currency"]
else:
    currency.value = "INR"

item_name = data["item_name"].split("#%#")
item_amount = data["item_amount"].split("#%#")

for i,s in enumerate(item_name):
    item_name[i]=s.replace("\n", " ")

item_igst_percent = []
item_cgst_percent = []
item_sgst_percent = []
item_hsn = []
item_rate = []
item_discount_percent = []

if "item_igst_percent" in data:
    item_igst_percent = data["item_igst_percent"].split("#%#")
if "item_cgst_percent" in data:
    item_cgst_percent = data["item_cgst_percent"].split("#%#")
if "item_sgst_percent" in data:
    item_sgst_percent = data["item_sgst_percent"].split("#%#")
if "item_hsn" in data:
    item_hsn = data["item_hsn"].split("#%#")
if "item_quantity" in data:
    item_quantity = data["item_quantity"].split("#%#")
else:
    item_quantity = [1]*len(item_name)
if "item_rate" in data:
    item_rate = data["item_rate"].split("#%#")
if "item_discount_percent" in data:
    item_discount_percent = data["item_discount_percent"].split("#%#")


for i, item in enumerate(item_name):
    n = str(i+18)
    sheet['a'+n].value = i+1
    product_id = sheet['b'+n]
    sku = sheet['c'+n]
    hsn = sheet['d'+n]
    title = sheet['e'+n]
    quantity = sheet['f'+n]
    unit_price = sheet['g'+n]
    excise_duty = sheet['h'+n]
    discount_percent = sheet['i'+n]
    sgst_percent = sheet['j'+n]
    cgst_percent = sheet['k'+n]
    igst_percent = sheet['l'+n]
    cess_percent = sheet['m'+n]
    tcs_percent = sheet['n'+n]
    total_amount = sheet['o'+n]

    title.value = item_name[i]
    quantity.value = item_quantity[i]
    total_amount.value = item_amount[i]

    if len(item_hsn):
        hsn.value = item_hsn[i]
    if len(item_igst_percent):
        igst_percent.value = item_igst_percent[i]
    if len(item_cgst_percent):
        cgst_percent.value = item_cgst_percent[i]
    if len(item_sgst_percent):
        sgst_percent.value = item_sgst_percent[i]
    if len(item_discount_percent):
        discount_percent.value = item_discount_percent[i]
    if len(item_rate):
        unit_price.value = item_rate[i]

    for cell in sheet[n]:
        cell.border = border

tab = Table(displayName="Table1", ref="A17:p"+str(18+i))
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=True)
tab.TableStyleInfo = style
ws.add_table(tab)

align = Alignment(horizontal = 'center')
sheet.merge_cells('b'+str(19+i)+':e'+str(19+i))
sheet['b'+str(19+i)].value = "Line Total"
sheet['b'+str(19+i)].alignment = align
sheet['b'+str(19+i)].font = Font(bold=True) 

net_total = sheet['o'+str(19+i)]
net_total.value = data["net_total"]

for cell in sheet[str(19+i)]:
    cell.border = border

wb.save(save_file)